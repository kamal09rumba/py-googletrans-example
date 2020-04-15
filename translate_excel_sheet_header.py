import openpyxl
from googletrans import Translator

file_name = "excel_sheet.xlsx"
translator = Translator()
# wb = openpyxl.load_workbook(file_name, read_only=True)
wb = openpyxl.load_workbook(file_name)
sheetnames = wb.sheetnames
for sheetname in sheetnames:
    active_sheet = wb[sheetname]
    sheetname_en = translator.translate(sheetname, dest="en").text.lower()
    print(sheetname, "-->", sheetname_en)
    print("+++++")
    active_sheet.title = sheetname_en
    col_no = active_sheet.max_column
    for i in range(col_no):
        value = active_sheet.cell(row=1, column=i + 1).value
        value_en = translator.translate(value, dest="en").text.lower()
        active_sheet.cell(row=1, column=i + 1, value=value_en)
        print(active_sheet.cell(row=1, column=i + 1).value, "-->", value_en)
    print()
    print("==========================================================")

wb.save("excel_sheet_en.xlsx")
